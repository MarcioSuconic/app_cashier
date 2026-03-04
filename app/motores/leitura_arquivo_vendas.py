class Leitura_Arquivo_Vendas:

    def __init__(self, data_sql):
        self.date_sql = data_sql
        self.__path_file_sales()
        self.__verify_file_exist()
        self.__follow_the_flow()

    def __path_file_sales(self):
        from app.utilitarios.arquivo_vendas.utils import return_read_path_file_sales
        self.path_file_sales = return_read_path_file_sales(self.date_sql)

    def __verify_file_exist(self):
        self.exist_file = self.path_file_sales.exists()

    def __follow_the_flow(self):

        if self.exist_file == True:
            self.__read_datas_sales()
            #self.save_datas_sales()

    def __read_datas_sales(self):        
        from openpyxl import load_workbook
        from decimal import Decimal

        file_sales = load_workbook(self.path_file_sales)
        aba_file_sales = file_sales['Completo']

        last_row = aba_file_sales.max_row

        modo_inserir = False

        self.total_sales_day = 0

        self.dict_geral = {}
        lista_dados = []

        for linha in range(8,last_row):
            first_collumn_ativo = aba_file_sales.cell(linha, 1).value

            if modo_inserir == True and first_collumn_ativo != None:
                codigo_barras = str(aba_file_sales.cell(linha, 1).value)
                qtde = int(aba_file_sales.cell(linha, 3).value)
                vl_unit = float(aba_file_sales.cell(linha, 5).value)
                vl_tot = float(aba_file_sales.cell(linha, 6).value)
                lista_temp = [codigo_barras,qtde,vl_unit,vl_tot]
                lista_dados.append(lista_temp)
                valor = float(aba_file_sales.cell(linha, 6).value)
                self.total_sales_day += valor

            if first_collumn_ativo == '1':
                modo_inserir = True
                n_boleto = aba_file_sales.cell(linha-1, 1).value

            if first_collumn_ativo == None and modo_inserir == True:
                dict_temp ={n_boleto:lista_dados}
                lista_dados = []
                self.dict_geral = {**self.dict_geral, **dict_temp}
                dict_temp = {}
            
            if first_collumn_ativo == None:
                modo_inserir = False

    def save_datas_sales(self):
        import json
        from app.utilitarios.arquivo_vendas.utils import return_save_path_file_sales
        file_path_save = return_save_path_file_sales(self.date_sql)

        with open(file_path_save, 'w', encoding='utf-8') as file:
            json.dump(self.dict_geral, file, ensure_ascii=False, indent=4)

    def return_total_sales_day(self):
        print(f'O total vendido em {self.date_sql} foi R$ {round(self.total_sales_day,2)}')
        return round(self.total_sales_day,2)

    def return_exist_file(self):
        return self.exist_file